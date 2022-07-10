import logging

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import time

class RoboRafael:
    def __init__(self):
        self.chrome = webdriver.Chrome()

    def correio(self):
        arq = "GoLiveTech-ExercicioRafael-BaseCEPs.xlsx"
        site_correio = "https://buscacepinter.correios.com.br/app/endereco/index.php"
        df = pd.read_excel(arq)
        wb = Workbook()
        ws = wb.active
        self.columns_excel(wb, ws)
        self.automate(df, site_correio, wb, ws)

    def columns_excel(self, wb, ws):
        ws['A1'] = 'Logradouro'
        ws['B1'] = 'Bairro'
        ws['C1'] = 'Localidade'
        ws['D1'] = 'CEP'
        wb.save('GoLiveTech-ExercicioRafael-NovaBaseCEPs.xlsx')

    def write_excel(self, wb, ws):
        logradouro = self.chrome.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[1]').text
        bairro = self.chrome.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[2]').text
        localidade = self.chrome.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[3]').text
        cep = self.chrome.find_element(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[4]').text

        to_append = [str(logradouro), str(bairro), str(localidade), str(cep)]
        ws.append(to_append)
        wb.save('GoLiveTech-ExercicioRafael-NovaBaseCEPs.xlsx')

    def automate(self, df, site_correio, wb, ws):
        for index, row in df.iterrows():
            value = row['CEP']
            value = value.replace("-", "")
            self.chrome = webdriver.Chrome(executable_path='/usr/local/bin/chromedriver')
            self.chrome.get(site_correio)

            element_cep = self.chrome.find_element(By.XPATH, '//*[@id="endereco"]')
            element_cep.send_keys(value)

            element_button = self.chrome.find_element(By.XPATH, '//*[@id="btn_pesquisar"]')
            element_button.click()
            time.sleep(3)
            self.write_excel(wb, ws)
